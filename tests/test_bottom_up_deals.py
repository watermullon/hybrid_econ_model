from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError

from src.config_loader import load_inputs
from src.deal_model import (
    calculate_asset_value,
    calculate_capex,
    calculate_debt_service,
    calculate_noi,
    calculate_refi_capacity,
    run_deal_year,
)
from src.deal_types import DealConfig, DealSet
from src.engine import run_scenario
from src.model_types import ModelConfig, Scenario, ScenarioSet
from src.outputs import write_outputs
from src.portfolio_aggregator import build_re_portfolio_years
from src.utils import merge_model


ROOT = Path(__file__).resolve().parents[1]


class NestedMergeChild(BaseModel):
    x: int = Field(ge=0)
    y: int = Field(ge=0)


class NestedMergeModel(BaseModel):
    child: NestedMergeChild
    sibling: str


def make_deal(**overrides) -> DealConfig:
    data = {
        "enabled": True,
        "acquisition_year": 1,
        "description": "Test deal",
        "acquisition": {
            "asset_value": 23_000_000,
            "value_type": "unknown",
            "purchase_price": None,
            "new_equity_required": 2_000_000,
            "closing_costs": 0,
            "initial_capex_reserve": 0,
        },
        "capital_stack": {
            "assumed_debt": 16_000_000,
            "assumed_liabilities": 1_000_000,
            "liabilities_written_down": 0,
            "seller_note": 0,
            "preferred_equity": 0,
        },
        "operations": {
            "current_noi": 1_200_000,
            "stabilized_noi": 1_800_000,
            "years_to_stabilization": 3,
            "annual_noi_growth_after_stabilization": 0.02,
            "gross_rent": None,
            "gross_rent_growth": 0.02,
            "vacancy_rate": None,
        },
        "debt": {
            "interest_rate": 0.065,
            "maturity_year": 5,
            "amortization_type": "interest_only",
            "annual_debt_service": None,
            "amortization_years": None,
            "dscr_minimum": 1.25,
            "recourse": "unknown",
        },
        "capex": {"annual_capex": {1: 300_000, 2: 250_000, 3: 150_000}, "recurring_capex_pct_of_noi": 0.03},
        "valuation": {"method": "growth", "exit_cap_rate": None, "annual_value_growth": 0.03, "stabilized_value": None},
        "refinance": {
            "enabled": True,
            "target_years": [5],
            "refi_ltv": 0.65,
            "refi_costs_pct": 0.02,
            "max_cash_out_pct_of_value": None,
            "proceeds_use": "fund_liquidity",
        },
    }
    for key, value in overrides.items():
        if isinstance(value, dict) and isinstance(data.get(key), dict):
            data[key] = {**data[key], **value}
        else:
            data[key] = value
    return DealConfig.model_validate(data)


def make_config(*, bottom_up: bool = False) -> ModelConfig:
    config, _, _ = load_inputs(ROOT / "inputs")
    if not bottom_up:
        return config
    return config.model_copy(update={"real_estate": config.real_estate.model_copy(update={"mode": "bottom_up"})})


def make_scenario(years: int = 5, **overrides) -> Scenario:
    data = {
        "description": "Bottom-up test scenario",
        "years": years,
        "real_estate": {
            "initial_noi_yield": 0.07,
            "annual_noi_growth": 0.02,
            "annual_nav_appreciation": 0.03,
            "gross_rent_yield": 0.11,
        },
        "hedge_fund": {"annual_returns": [0.0] * years},
    }
    data.update(overrides)
    return Scenario.model_validate(data)


def test_deep_merge_preserves_nested_siblings_and_original_model() -> None:
    base = NestedMergeModel(child=NestedMergeChild(x=1, y=2), sibling="keep")

    merged = merge_model(base, {"child": {"x": 9}})

    assert merged.child.x == 9
    assert merged.child.y == 2
    assert merged.sibling == "keep"
    assert base.child.x == 1


def test_deep_merge_revalidates_nested_override() -> None:
    base = NestedMergeModel(child=NestedMergeChild(x=1, y=2), sibling="keep")

    with pytest.raises(ValidationError):
        merge_model(base, {"child": {"x": -1}})


def test_deal_model_calculates_nav_cashflow_dscr_and_refi_timing() -> None:
    deal = make_deal()

    row = run_deal_year(scenario_name="s", deal_name="d", deal=deal, model_year=1)

    assert row.deal_nav == pytest.approx(6_000_000)
    assert row.entry_equity_cushion == pytest.approx(4_000_000)
    assert calculate_debt_service(deal, 16_000_000) == pytest.approx(1_040_000)
    assert row.debt_service == pytest.approx(1_040_000)
    assert row.capex == pytest.approx(336_000)
    assert row.free_cashflow_after_debt_and_capex == pytest.approx(-176_000)
    assert row.dscr == pytest.approx(1_200_000 / 1_040_000)
    assert row.value_to_new_equity_multiple == pytest.approx(3.0)
    assert row.refi_proceeds == 0

    refi_row = run_deal_year(scenario_name="s", deal_name="d", deal=deal, model_year=5)
    assert refi_row.refi_proceeds == pytest.approx(refi_row.refi_capacity)
    assert refi_row.refi_liability_added == pytest.approx(refi_row.refi_proceeds)


def test_deal_helpers_cover_noi_interpolation_fixed_debt_cap_rate_and_refi_floor() -> None:
    fixed_debt_deal = make_deal(
        debt={"amortization_type": "fixed_annual_debt_service", "annual_debt_service": 900_000, "interest_rate": None}
    )
    cap_rate_deal = make_deal(valuation={"method": "cap_rate", "exit_cap_rate": 0.08})
    underwater_deal = make_deal(acquisition={"asset_value": 10_000_000}, capital_stack={"assumed_debt": 16_000_000})

    assert calculate_noi(fixed_debt_deal, 1) == pytest.approx(1_200_000)
    assert calculate_noi(fixed_debt_deal, 2) == pytest.approx(1_500_000)
    assert calculate_noi(fixed_debt_deal, 4) == pytest.approx(1_800_000 * 1.02)
    assert calculate_debt_service(fixed_debt_deal, 16_000_000) == pytest.approx(900_000)
    assert calculate_capex(fixed_debt_deal, 2, 1_500_000) == pytest.approx(295_000)
    assert calculate_asset_value(fixed_debt_deal, 2, 1_500_000) == pytest.approx(23_000_000 * 1.03)
    assert calculate_asset_value(cap_rate_deal, 2, 1_500_000) == pytest.approx(18_750_000)
    assert calculate_refi_capacity(underwater_deal, 10_000_000, 16_000_000) == 0


def test_portfolio_aggregation_handles_inactive_active_multiple_deals_and_overrides() -> None:
    deal_one = make_deal()
    deal_two = make_deal(acquisition_year=2, acquisition={"asset_value": 5_000_000, "new_equity_required": 1_000_000}, capital_stack={"assumed_debt": 2_000_000, "assumed_liabilities": 0})
    deals = DealSet(deals={"one": deal_one, "two": deal_two})

    portfolio, deal_rows = build_re_portfolio_years(
        scenario_name="s",
        years=2,
        deals=deals,
        deal_overrides={"one": {"operations": {"current_noi": 1_300_000}}},
    )

    assert portfolio[0].active_deal_count == 1
    assert [row.active for row in deal_rows if row.deal_name == "two" and row.year == 1] == [False]
    assert portfolio[1].active_deal_count == 2
    assert portfolio[0].noi == pytest.approx(1_300_000)
    assert portfolio[1].gross_asset_value == pytest.approx(23_000_000 * 1.03 + 5_000_000)


def test_engine_top_down_default_is_unchanged_when_deals_are_loaded() -> None:
    config, scenario_set, deals = load_inputs(ROOT / "inputs")
    scenario = scenario_set.scenarios["base_hit_everyone_happy"]

    without_deals = run_scenario("base", scenario, config)
    with_deals = run_scenario("base", scenario, config, deals)

    assert with_deals.summary["real_estate_mode"] == "top_down"
    assert with_deals.cashflows[0].re_noi == pytest.approx(without_deals.cashflows[0].re_noi)
    assert with_deals.deal_cashflows == []


def test_engine_bottom_up_requires_deals() -> None:
    config = make_config(bottom_up=True)
    scenario = make_scenario()

    with pytest.raises(ValueError, match="no deals"):
        run_scenario("bottom", scenario, config, None)


def test_engine_bottom_up_uses_deal_nav_cashflow_and_refi_proceeds() -> None:
    config = make_config(bottom_up=True)
    deal = make_deal(
        capital_stack={"assumed_debt": 5_000_000},
        refinance={"target_years": [1], "proceeds_use": "reserve"},
    )
    scenario = make_scenario(years=1)

    result = run_scenario("bottom", scenario, config, DealSet(deals={"deal": deal}))
    row = result.cashflows[0]

    assert row.real_estate_mode == "bottom_up"
    assert row.re_closing_nav == pytest.approx(17_000_000)
    assert row.re_free_cashflow_after_debt_and_capex == pytest.approx(539_000)
    assert row.lp_distribution >= 0
    assert row.re_refi_proceeds_from_deals > 0
    assert row.refinance_liability == pytest.approx(row.re_refi_proceeds_from_deals)
    assert row.reserve_closing_nav > 8_000_000


def test_engine_bottom_up_negative_cashflow_uses_cash_sources_before_shortfall() -> None:
    config = make_config(bottom_up=True)
    deal = make_deal(
        operations={"current_noi": 100_000, "stabilized_noi": 100_000},
        capex={"annual_capex": {1: 9_500_000}, "recurring_capex_pct_of_noi": 0.0},
    )
    scenario = make_scenario(years=1)

    result = run_scenario("bottom", scenario, config, DealSet(deals={"deal": deal}))
    row = result.cashflows[0]

    assert row.lp_distribution >= 0
    assert row.re_cashflow_shortfall > 0
    assert "RE_CASHFLOW_SHORTFALL" in row.event_flag


def test_engine_bottom_up_hurdle_trigger_can_execute() -> None:
    config = make_config(bottom_up=True).model_copy(
        update={
            "waterfall": make_config().waterfall.model_copy(update={"lp_hurdle_moic": 1.1}),
            "hurdle_completion_trigger": make_config().hurdle_completion_trigger.model_copy(
                update={"minimum_lp_cash_moic_before_trigger": 0.0}
            ),
            "backend_liquidity_strategy": make_config().backend_liquidity_strategy.model_copy(
                update={"enabled": False}
            ),
        }
    )
    deal = make_deal(acquisition={"asset_value": 30_000_000}, capital_stack={"assumed_debt": 5_000_000})
    scenario = make_scenario(years=3)

    result = run_scenario("bottom", scenario, config, DealSet(deals={"deal": deal}))

    assert result.summary["lp_hurdle_achieved"] is True
    assert result.summary["hurdle_trigger_executed"] is True


def test_outputs_include_bottom_up_columns_and_deal_cashflow_sheet(tmp_path: Path) -> None:
    config = make_config(bottom_up=True).model_copy(
        update={"reporting": make_config().reporting.model_copy(update={"output_csv": True, "output_excel": True})}
    )
    deal = make_deal()
    scenario = make_scenario(years=1)
    scenario_set = ScenarioSet(scenarios={"bottom": scenario})
    result = run_scenario("bottom", scenario, config, DealSet(deals={"deal": deal}))

    frames = write_outputs(results=[result], config=config, scenarios=scenario_set, output_dir=tmp_path)

    assert "total_deal_refi_proceeds" in frames["summary"].columns
    assert "re_gross_asset_value" in frames["cashflows"].columns
    assert (tmp_path / "deal_cashflows.csv").exists()
    assert not pd.read_csv(tmp_path / "deal_cashflows.csv").empty
    workbook = load_workbook(tmp_path / "scenario_summary.xlsx", read_only=True)
    assert "Deal Cashflows" in workbook.sheetnames
